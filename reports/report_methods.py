import pandas as pd
from openpyxl import load_workbook
from reports.report_utils import apply_border,column_widths,highlight_cells,has_not_expected_or_missed,has_not_expected,is_positive_number,is_zero
from reports.styles import red_fill, green_fill, bold_font

from utils.colors import Colors


def total_report(records, output_file):
    try:
        df = pd.DataFrame(records)
        df.to_excel(output_file, index=False)
        wb = load_workbook(output_file)
        ws = wb.active

        column_widths(ws)
        apply_border(ws)
        
        highlight_cells(ws, (2, 2), has_not_expected_or_missed, red_fill, None)
        highlight_cells(ws, (1, 1), has_not_expected, red_fill, None)
        highlight_cells(ws, (6, 6), is_positive_number, green_fill, red_fill)
        highlight_cells(ws, (7, 10), is_zero, green_fill, red_fill)

        for cell in ws[1]:
            cell.font = bold_font

        wb.save(output_file)
        Colors.colored_print(f'TOTAL REPORT saved at: {output_file}', 'OKGREEN')

    except Exception as e:
        Colors.colored_print(f'ERROR creating TOTAL REPORT: {e}', "FAIL")
        # print(f'{Colors.FAIL} ERROR creating TOTAL REPORT: {e}{Colors.ENDC}')

def difference_report(records, output_file, decimal):
    try:
        df = pd.DataFrame(records)
        df.to_excel(output_file, index=False)
        wb = load_workbook(output_file)
        ws = wb.active

        column_widths(ws)
        apply_border(ws)

        highlight_cells(ws, (5, 5),
                        lambda v: isinstance(v, (int, float)) and abs(v) > 10 ** -decimal,
                        # lambda v: v > 0, red_fill)

                        red_fill, green_fill)

        wb.save(output_file)
        Colors.colored_print(f'DIFFERENCE REPORT saved at: {output_file}', 'OKGREEN')
    except Exception as e:
        Colors.colored_print(f'ERROR creating DIFFERENCE REPORT: {e}', "FAIL")

def highlighted_report(records, output_file):
    try:
        df = pd.DataFrame(records)
        df.to_excel(output_file, index=False)
        wb = load_workbook(output_file)
        ws = wb.active

        column_widths(ws)
        apply_border(ws)

        highlight_cells(ws, (4, 4),
                        lambda v: v == "PASS",
                        green_fill, red_fill)

        wb.save(output_file)
        Colors.colored_print(f'HIGHLIGHTED REPORT saved at: {output_file}', 'OKGREEN')
    except Exception as e:
        Colors.colored_print(f'ERROR creating HIGHLIGHTED REPORT: {e}', "FAIL")
        
        
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.styles import Border, Side, PatternFill, Font
# from openpyxl.utils import get_column_letter
# from utils.colors import Colors

# # === Styles ===
# thin_border = Border(
#     left=Side(style="thin"),
#     right=Side(style="thin"),
#     top=Side(style="thin"),
#     bottom=Side(style="thin")
# )

# green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
# red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
# bold_font = Font(bold=True)

# # === Helper functions ===
# def apply_thin_border(ws):
#     for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#         for cell in row:
#             cell.border = thin_border

# def auto_adjust_column_widths(ws):
#     for col in ws.columns:
#         max_length = 0
#         col_letter = get_column_letter(col[0].column)
#         for cell in col:
#             if cell.value:
#                 max_length = max(max_length, len(str(cell.value)))
#         ws.column_dimensions[col_letter].width = max_length + 2

# def highlight_cells(ws, col_range, condition_fn, fill_true, fill_false):
#     for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_range[0], max_col=col_range[1]):
#         for cell in row:
#             if condition_fn(cell.value):
#                 cell.fill = fill_true
#             else:
#                 cell.fill = fill_false

# # === Report functions ===
# def total_report(records, output_file):
#     try:
#         df = pd.DataFrame(records)
#         df.to_excel(output_file, index=False)
#         wb = load_workbook(output_file)
#         ws = wb.active

#         auto_adjust_column_widths(ws)
#         apply_thin_border(ws)

#         # Highlight column 2 if value contains "NOT EXPECTED" or "missed"
#         highlight_cells(ws, (2, 2), 
#                         lambda v: isinstance(v, str) and ("NOT EXPECTED" in v or "missed" in v),
#                         red_fill, None)

#         # Highlight column 1 similarly
#         highlight_cells(ws, (1, 1),
#                         lambda v: isinstance(v, str) and "NOT EXPECTED" in v,
#                         red_fill, None)

#         # Highlight column 6 based on positive/zero
#         highlight_cells(ws, (6, 6),
#                         lambda v: isinstance(v, (int, float)) and v > 0,
#                         green_fill, red_fill)

#         # Highlight columns 7 to 10 based on value being 0 or not
#         highlight_cells(ws, (7, 10),
#                         lambda v: v == 0,
#                         green_fill, red_fill)

#         # Bold header row
#         for cell in ws[1]:
#             cell.font = bold_font

#         wb.save(output_file)
#         Colors.colored_print(f'TOTAL REPORT saved at: {output_file}', 'OKGREEN')

#     except Exception as e:
#         Colors.colored_print(f'ERROR creating TOTAL REPORT: {e}', "FAIL")

# def difference_report(records, output_file, decimal=4):
#     try:
#         df = pd.DataFrame(records)
#         df.to_excel(output_file, index=False)
#         wb = load_workbook(output_file)
#         ws = wb.active

#         auto_adjust_column_widths(ws)
#         apply_thin_border(ws)

#         # Highlight column 5 based on difference threshold
#         highlight_cells(ws, (5, 5),
#                         lambda v: isinstance(v, (int, float)) and abs(v) > 10 ** -decimal,
#                         red_fill, green_fill)

#         wb.save(output_file)
#         Colors.colored_print(f'DIFFERENCE REPORT saved at: {output_file}', 'OKGREEN')
#     except Exception as e:
#         Colors.colored_print(f'ERROR creating DIFFERENCE REPORT: {e}', "FAIL")

# def highlighted_report(records, output_file):
#     try:
#         df = pd.DataFrame(records)
#         df.to_excel(output_file, index=False)
#         wb = load_workbook(output_file)
#         ws = wb.active

#         auto_adjust_column_widths(ws)
#         apply_thin_border(ws)

#         # Highlight column 4 based on PASS/FAIL
#         highlight_cells(ws, (4, 4),
#                         lambda v: v == "PASS",
#                         green_fill, red_fill)

#         wb.save(output_file)
#         Colors.colored_print(f'HIGHLIGHTED REPORT saved at: {output_file}', 'OKGREEN')
#     except Exception as e:
#         Colors.colored_print(f'ERROR creating HIGHLIGHTED REPORT: {e}', "FAIL")