from reports.styles import thin_border
from openpyxl.utils import get_column_letter

def apply_border(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

def column_widths(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

# def highlight_cells(ws, col_range, condition_fn, fill_true, fill_false):
#     for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_range[0], max_col=col_range[1]):
#         for cell in row:
#             if condition_fn(cell.value):
#                 cell.fill = fill_true
#             else:
#                 cell.fill = fill_false
def highlight_cells(ws, col_range, condition_fn, fill_true, fill_false):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_range[0], max_col=col_range[1]):
        for cell in row:
            if condition_fn(cell.value):
                cell.fill = fill_true
            elif fill_false is not None:  
                cell.fill = fill_false                
def has_not_expected_or_missed(obj):
    return isinstance(obj, str) and ("NOT EXPECTED" in obj or "missed" in obj)

def has_not_expected(obj):
    return isinstance(obj, str) and "NOT EXPECTED" in obj

def is_positive_number(obj):
    return isinstance(obj, (int, float)) and obj > 0

def is_zero(obj):
    return obj == 0


__all__ = ['is_zero', 'is_positive_number',
           'has_not_expected', 'has_not_expected_or_missed',
           'highlight_cells', 'column_widths',
           'apply_border']