
import pandas as pd
from openpyxl import load_workbook
from comparator_app.reports.report_utils import apply_border, column_widths, highlight_cells, has_not_expected_or_missed, has_not_expected, is_positive_number, is_zero,red_fill, green_fill, bold_font

# from reports.report_utils import apply_border, column_widths, highlight_cells, has_not_expected_or_missed, has_not_expected, is_positive_number, is_zero,red_fill, green_fill, bold_font
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter



def highlighted_report(file_name, sheet_name, records, output_file):
    # decimal = config.get('decimal', 5)
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = ["Key", "Value File 1", "Value File 2", "Match", "Max_difference", "Difference"]
    ws.append(headers)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for record in records:
        key_raw = record.get('key', "")
    
        if isinstance(key_raw, str) or "missing" in str(key_raw).lower():
            key = str(key_raw)
        else:
            key = " | ".join(key_raw)
        value_file1 = str(record.get('value_file1', []))
        value_file2 = str(record.get('value_file2', []))
        match = record.get('match', "FAIL")
        difference = str(record.get('difference', []))
        max_diff = record.get('max_diff', "")
        
        row = [key, value_file1, value_file2, match, max_diff, difference]
        ws.append(row)
        
        fill = None
        try:
            if isinstance(max_diff, (int, float)) and abs(max_diff) > 10 ** -decimal:
                fill = fail_fill
        except (TypeError, ValueError, NameError) as e:
            print(f"Warning: max_diff check failed - {e}")

        if fill:
            for col in range(1, len(headers)+1):
                ws.cell(row=ws.max_row, column=col).fill = fill

        fill = pass_fill if match == "PASS" else fail_fill
        for col in range(1, len(headers)+1):
            ws.cell(row=ws.max_row, column=col).fill = fill
            
        apply_border(ws)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    wb.save(output_file)

def total_report(records, output_file):
    generate_excel_report(
        records=records,
        output_file=output_file,
        bold_header=True,
        highlight_rules=[
            {
                "coords": (2, 2),
                "condition": has_not_expected_or_missed,
                "fill_pass": red_fill
            },
            {
                "coords": (1, 1),
                "condition": has_not_expected,
                "fill_pass": red_fill
            },
            {
                "coords": (6, 6),
                "condition": is_positive_number,
                "fill_pass": green_fill,
                "fill_fail": red_fill
            },
            {
                "coords": (7, 10),
                "condition": is_zero,
                "fill_pass": green_fill,
                "fill_fail": red_fill
            }
        ]
    )


def generate_excel_report(records, output_file: str, highlight_rules: list = None, bold_header: bool = False):
    try:
        df = pd.DataFrame(records)
        df.to_excel(output_file, index=False)

        wb = load_workbook(output_file)
        ws = wb.active

        column_widths(ws)
        apply_border(ws)
        col_name_to_idx = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

        if highlight_rules:
            for rule in highlight_rules:
                column = rule.get("column")
                coords = rule.get("coords")
                condition = rule["condition"]
                fill_pass = rule.get("fill_pass")
                fill_fail = rule.get("fill_fail")

                if column:
                    if column not in col_name_to_idx:
                        continue
                    col_idx = col_name_to_idx[column]
                    highlight_cells(ws, (col_idx, col_idx), condition, fill_pass, fill_fail)
                elif coords:
                    highlight_cells(ws, coords, condition, fill_pass, fill_fail)

        if bold_header:
            for cell in ws[1]:
                cell.font = bold_font

        wb.save(output_file)
        print(f'REPORT saved at: {output_file}', 'OKGREEN')

    except Exception as e:
        print(f'ERROR creating REPORT: {e}', "FAIL")


