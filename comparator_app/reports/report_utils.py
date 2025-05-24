import io
from flask import send_file
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from openpyxl.styles import Border, Side, PatternFill, Font


thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
green_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

bold_font = Font(bold=True)

def apply_border(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border


def column_widths(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2


def highlight_cells(ws, coords, condition, fill_pass, fill_fail):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row-1, min_col=coords[0], max_col=coords[1]):
        for cell in row:
            if condition(cell.value):
                cell.fill = fill_pass
            elif fill_fail is not None:
                cell.fill = fill_fail


def has_not_expected_or_missed(obj):
    return isinstance(obj, str) and ("NOT EXPECTED" in obj or "missed" in obj)


def has_not_expected(obj):
    return isinstance(obj, str) and "NOT EXPECTED" in obj


def is_positive_number(obj):
    return isinstance(obj, (int, float)) and obj > 0


def is_zero(obj):
    return obj == 0


def format_and_send_excel(
    df,
    sheet_name: str,
    download_name: str,
    highlight_rules: list = None,
    bold_header: bool = False
):
    try:
        output = io.BytesIO()
        df.to_excel(output, index=False, sheet_name=sheet_name)
        output.seek(0)

        wb = load_workbook(output)
        ws = wb[sheet_name]

        column_widths(ws)
        apply_border(ws)

        if highlight_rules:
            for rule in highlight_rules:
                highlight_cells(
                    ws,
                    coords=rule["coords"],
                    condition=rule["condition"],
                    fill_pass=rule.get("fill_pass"),
                    fill_fail=rule.get("fill_fail")
                )

        if bold_header:
            for cell in ws[1]:
                cell.font = bold_font

        new_output = io.BytesIO()
        wb.save(new_output)
        new_output.seek(0)

        return send_file(
            new_output,
            download_name=download_name,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f'ERROR generating downloadable report: {e}')
        return f"❌ Error creating report: {e}", 500


__all__ = [
    "format_and_send_excel",
    "highlight_cells",
    "column_widths",
    "apply_border",
    "is_zero",
    "is_positive_number",
    "has_not_expected",
    "has_not_expected_or_missed"
]


