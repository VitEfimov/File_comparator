from openpyxl.styles import Border, Side, PatternFill, Font


thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
# red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# green_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
# red_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
green_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

bold_font = Font(bold=True)